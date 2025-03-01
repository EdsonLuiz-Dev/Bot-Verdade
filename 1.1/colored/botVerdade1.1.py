from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from colored import Fore
from math import pow

def gerarTabela(numEntradas):
    tabelaResultado = []
    linhas = pow(2, numEntradas)
    loop_range = 1

    for j in range(0, numEntradas):
        temp = []
        linhas = linhas / 2
        for k in range(0, loop_range):
            for i in range(0, int(linhas)): 
                temp.append("1")
            for i in range(0, int(linhas)):
                temp.append("0")
        loop_range = loop_range * 2
        tabelaResultado.append(temp)

    return tabelaResultado


def gerarTabelaVerdadeiroFalso(numEntradas):
    tabelaResultado = []
    linhas = pow(2, numEntradas)  
    loop_range = 1

    for j in range(0, numEntradas):
        temp = []
        linhas = linhas / 2
        for k in range(0, loop_range):
            for i in range(0, int(linhas)): 
                temp.append("V")
            for i in range(0, int(linhas)):
                temp.append("F")
        loop_range = loop_range * 2
        tabelaResultado.append(temp)

    return tabelaResultado


def gerarTabelaInvertida(numEntradas):
    tabelaResultado = []
    linhas = pow(2, numEntradas)
    loop_range = 1

    for j in range(0, numEntradas):
        temp = []
        loop_range = loop_range * 2
        for i in range(0, int(linhas/loop_range)):
            for k in range(int(loop_range/2)):
                temp.append("1")
            for k in range(int(loop_range/2)):
                temp.append("0")
        tabelaResultado.append(temp)
    
    return tabelaResultado


def gerarTabelaInvertidaVerdadeiroFalso(numEntradas):
    tabelaResultado = []
    linhas = pow(2, numEntradas)
    loop_range = 1

    for j in range(0, numEntradas):
        temp = []
        loop_range = loop_range * 2
        for i in range(0, int(linhas/loop_range)):
            for k in range(int(loop_range/2)):
                temp.append("V")
            for k in range(int(loop_range/2)):
                temp.append("F")
        tabelaResultado.append(temp)
    
    return tabelaResultado


def injetarValores(numEntradas, worksheet, tabela):
    totLinhas = pow(2,numEntradas)
    for colunas in range(1, numEntradas+1):
        char = get_column_letter(colunas)
        worksheet[(char + "1")].value = char
        for linhas in range(2, int(totLinhas + 2)):
            worksheet[(char + str(linhas))].value = int(tabela[colunas-1][linhas-2])
    
    print("Valores injetados!")


def injetarValoresVerdadeiroFalso(numEntradas, worksheet, arrayValores):
    totLinhas = pow(2,numEntradas)
    for colunas in range(1, numEntradas+1):
        char = get_column_letter(colunas)
        worksheet[(char + "1")].value = char
        for linhas in range(2, int(totLinhas + 2)):
            worksheet[(char + str(linhas))].value = str(arrayValores[colunas-1][linhas-2])
    
    print("Valores injetados!")


autorizado = False
while not autorizado:
    print(f'Deseja {Fore.blue}CRIAR{Fore.white} ou {Fore.blue}CARREGAR{Fore.white} um arquivo?')
    print(f'{Fore.green}CRIAR - 1{Fore.white} // {Fore.green}CARREGAR - 2{Fore.white}')
    escolhaArquivo = int(input())

    if(escolhaArquivo == 1):
        filename = str(input(f'Digite o nome do arquivo a ser criado {Fore.red}(SEM A EXTENSÃO){Fore.white} : ')) + ".xlsx"
        wb = Workbook()
        autorizado = True
    elif(escolhaArquivo == 2):
        filename = str(input(f'Digite o nome do arquivo a ser carregado {Fore.red}(DEVE ESTAR NO MESMO DIRETÓRIO DO SCRIPT)(SEM A EXTENSÃO){Fore.white} : ')) + ".xlsx"
        wb = load_workbook(filename)
        autorizado = True
    else:
        print(f'{Fore.red}Escolha uma opção VÁLIDA{Fore.white}')

autorizado = False
while not autorizado:
    print(f'Você deseja criar uma tabela {Fore.blue}INVERTIDA{Fore.white} ou em {Fore.blue}ORDEM ALFABÉTICA?{Fore.white}')
    print(f'{Fore.green}INVERTIDA - 1{Fore.white} // {Fore.green}ORDEM ALFABÉTICA - 2{Fore.white}')
    escolhaOrdem = int(input())

    if(escolhaOrdem == 1 or escolhaOrdem == 2):
        autorizado = True
    else:
        print(f'{Fore.red}Escolha uma opção VÁLIDA{Fore.white}')

autorizado = False
while not autorizado:
    print(f'Você deseja usar {Fore.blue}ZEROS e UNS{Fore.white} ou {Fore.blue}"V" e "F"{Fore.white}?')
    print(f'{Fore.green}ZEROS E UNS{Fore.white} - 1 // {Fore.green}"V" e "F"{Fore.white} - 2')
    escolhaCaractere = int(input())

    if(escolhaCaractere == 1 or escolhaCaractere == 2):
        autorizado = True
    else:
        print(f'{Fore.red}Escolha uma opção VÁLIDA{Fore.white}')

ws = wb.active
entradas = int(input("Digite o número de entradas desejadas: "))

if(escolhaCaractere == 1):
    if(escolhaOrdem == 1):
        tabela = gerarTabelaInvertida(entradas)
    else:
        tabela = gerarTabela(entradas)
else:
    if(escolhaOrdem == 1):
        tabela = gerarTabelaInvertidaVerdadeiroFalso(entradas)
    else:
        tabela = gerarTabelaVerdadeiroFalso(entradas)

if(escolhaCaractere == 1):
    injetarValores(entradas, ws, tabela)
else:
    injetarValoresVerdadeiroFalso(entradas, ws, tabela)


wb.save(filename)