from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from math import pow

selecionado = False

while not selecionado:
    print("Deseja CRIAR um arquivo ou CARREGAR um arquivo?")
    print("CRIAR - 1 // CARREGAR - 2")
    escolha = (int(input("Opção: ")))
    if(escolha == 1):
        filename = str(input("Informe o nome do arquivo a ser criado: ")) + ".xlsx"
        selecionado = True
        wb = Workbook()
    elif(escolha == 2):
        filename = str(input("Digite o nome do arquivo a ser carregado (DEVE ESTAR NO MESMO DIRETÓRIO DO SCRIPT): ")) + ".xlsx"
        selecionado = True
        wb = load_workbook(filename)
    
ws = wb.active

entradas = int(input("Digite o número de entradas: "))
linhas = int(pow(2, entradas))
temp = []
columnsArray = []
max_range = 1

for i in range(0, entradas):
    max_range = (max_range * 2)
    for j in range(0, int(linhas/max_range)):
        for k in range(0, int(max_range/2)):
            temp.append("1")
        for k in range(0, int(max_range/2)):
            temp.append("0")
    columnsArray.append(temp)
    temp = []

for col in range(1, entradas+1):
    char = get_column_letter(col)
    ws[(char + "1")].value = char;
    for row in range(2, linhas+2):
        ws[(char + str(row))].value = int(columnsArray[col-1][row-2])
                
wb.save(filename)