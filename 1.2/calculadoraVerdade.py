import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from math import pow

def gerarTabelaInvertida(numEntradas):
    tabelaResultado = []
    linhas = pow(2, numEntradas)
    loop_range = 1 

    for j in range(0, numEntradas):
        temp = []
        linhas = linhas / 2
        for k in range(0, loop_range):
            for i in range(0, int(linhas)): 
                temp.append("1")
            for i in range(0, int(linhas)):
                temp.append("0")
        loop_range = loop_range * 2
        tabelaResultado.append(temp)

    return tabelaResultado

def gerarTabela(numEntradas):
    tabelaResultado = []
    linhas = pow(2, numEntradas)
    loop_range = 1

    for j in range(0, numEntradas):
        temp = []
        loop_range = loop_range * 2
        for i in range(0, int(linhas/loop_range)):
            for k in range(int(loop_range/2)):
                temp.append("1")
            for k in range(int(loop_range/2)):
                temp.append("0")
        tabelaResultado.append(temp)
    
    return tabelaResultado

def injetarValores(numEntradas, worksheet, tabela):
    totLinhas = pow(2,numEntradas)
    for colunas in range(1, numEntradas+1):
        char = get_column_letter(colunas)
        worksheet[(char + "1")].value = char
        for linhas in range(2, int(totLinhas + 2)):
            worksheet[(char + str(linhas))].value = int(tabela[colunas-1][linhas-2])
    
    print("Valores injetados!")

def carregarColunas(coluna, linhas, ws):
    temp = []
    for i in range(2, int(linhas + 2)):
        temp.append(ws[coluna + str(i)].value)
    return temp

def nomeiaColuna(col, col2, op, ws):
    if(op == 1):
        return "(" + ws[col + '1'].value + " * " + ws[col2 + '1'].value + ")"
    elif(op == 2):
        return "(" + ws[col + '1'].value + " + " + ws[col2 + '1'].value + ")"
    elif(op == 3):
        return "(" + ws[col + '1'].value + " xor " + ws[col2 + '1'].value + ")"

def nomeiaColunaNot(col, ws):
    return ws[col + '1'].value + "'"

def inserirNovaColuna(colunaAInserir, contador, totLinhas, ws, col, col2):
    for coluna in range(1, entradas + contador):
        char = get_column_letter(coluna)
    for linhas in range(2, totLinhas + 2):
        ws[char + str(linhas)].value = colunaAInserir[linhas-2]
    ws[char + "1"].value = nomeiaColuna(col, col2, op, ws)

def inserirNovaColunaNot(colunaAInserir, contador, totLinhas, ws, col):
    for coluna in range(1, entradas + contador):
        char = get_column_letter(coluna)
    for linhas in range(2, totLinhas + 2):
        ws[char + str(linhas)].value = colunaAInserir[linhas-2]
    ws[char + "1"].value = nomeiaColunaNot(col, ws)

def AND(array1, array2):
    resultado = []
    for i in range(0, len(array1)):
        if(array1[i] * array2[i] == 0):
            resultado.append(0)
        else:
            resultado.append(1)
    return resultado

def OR(array1, array2):
    resultado = []
    for i in range(0, len(array1)):
        if(array1[i] + array2[i] >= 1):
            resultado.append(1)
        else:
            resultado.append(0)
    return resultado

def XOR(array1, array2):
    resultado = []
    for i in range(0, len(array1)):
        if((array1[i] + array2[i]) % 2 == 0):
            resultado.append(0)
        else:
            resultado.append(1)
    return resultado

def NOT(array):
    resultado = []
    for i in range(0, len(array)):
        if(array[i] == 1):
            resultado.append(0)
        else:
            resultado.append(1)
    return resultado

# Manipulação do arquivo
autorizado = False
while not autorizado:
    os.system('cls')
    print("========== CALCULADORA VERDADE ==========")
    print("Selecione as opções abaixo")
    print("CRIAR - 1 // CARREGAR - 2")
    escolhaArquivo = int(input("CRIAR ou CARREGAR? : "))
    if(escolhaArquivo == 1):
        filename = str(input("Digite o nome do arquivo a ser criado (SEM A EXTENSÃO) : ")) + ".xlsx"
        wb = Workbook()
        wb.save(filename)
        autorizado = True
        os.system('cls')
    elif(escolhaArquivo == 2):
        filename = str(input("Digite o nome do arquivo a ser carregado (DEVE ESTAR NO MESMO DIRETÓRIO DO SCRIPT)(SEM A EXTENSÃO) : ")) + ".xlsx"
        wb = load_workbook(filename)
        autorizado = True
        os.system('cls')
    else:
        print("Escolha uma opção VÁLIDA")

# Geração da tabela
autorizado = False
while not autorizado:
    print("ORDEM ALFABÉTICA - 1 // INVERTIDO - 2")
    escolhaOrdem = int(input("ORDEM ALFABÉTICA ou INVERTIDO? : "))

    if(escolhaOrdem == 1 or escolhaOrdem == 2):
        autorizado = True
    else:
        print("Escolha uma opção VÁLIDA.")
entradas = int(input("Digite o número de entradas desejadas: "))
os.system('cls')

ws = wb.active

# Injeção dos valores
if(escolhaOrdem == 1):
    tabela = gerarTabela(entradas)
else:
    tabela = gerarTabelaInvertida(entradas)
injetarValores(entradas, ws, tabela)

# Declaração de variáveis
linhas = int(pow(2, entradas))
contador = 1

# Loop que exibe a tabela no terminal e permite operar quantas vezes necessário
sair = True
while sair:
    # Exibe as colunas do arquivo
    print("Coluna Excel --> Operação/Entrada")
    for i in range(1, entradas + contador):
        char = get_column_letter(i)
        print(char + " --> " + str(ws[char + "1"].value))

    print("Digite a operação a ser realizada: ")
    print("AND - 1 // OR - 2 // XOR - 3 // NOT - 4 // FINALIZAR PROGRAMA - 0")
    op =  int(input())
    if(op == 0): # Verifica se o usuário não deseja finalizar o programa
        os.system('cls')
        print("Programa finalizado.")
        sair = False
        break

    if(op != 4): # Verifica se a operação escolhida não é uma negação
        col = str(input("Primeira Coluna : "))
        col2 = str(input("Segunda Coluna : "))
    else:
        col = str(input("Coluna a ser negada : "))
    
    # Realiza a operação escolhida e insere a nova coluna
    match op:
        case 1: # And    
            operando1 = carregarColunas(col, linhas, ws)
            operando2 = carregarColunas(col2, linhas, ws)
            resultado = AND(operando1, operando2)
            contador += 1
            inserirNovaColuna(resultado, contador, linhas, ws, col, col2)
            wb.save(filename)
            os.system('cls')
        case 2: # Or 
            operando1 = carregarColunas(col, linhas, ws)
            operando2 = carregarColunas(col2, linhas, ws)
            resultado = OR(operando1, operando2)
            contador += 1
            inserirNovaColuna(resultado, contador, linhas, ws, col, col2)
            wb.save(filename)
            os.system('cls')
        case 3: # Xor
            operando1 = carregarColunas(col, linhas, ws)
            operando2 = carregarColunas(col2, linhas, ws)
            resultado = XOR(operando1, operando2)
            contador += 1
            inserirNovaColuna(resultado, contador, linhas, ws, col, col2)
            wb.save(filename)
            os.system('cls')
        case 4: # Not 
            operando = carregarColunas(col, linhas, ws)
            resultado = NOT(operando)
            contador += 1
            inserirNovaColunaNot(resultado, contador, linhas, ws, col)
            wb.save(filename)
            os.system('cls')
        case _:
            print("Seleção inválida")
            os.system('cls')

wb.save(filename)